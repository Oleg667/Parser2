#https://metanit.com/python/tkinter/2.9.php

import tkinter

import pandas as pd
import json

import openpyxl
import time
import re





from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl.utils import get_column_letter

from tkinter import *
from tkinter.messagebox import showinfo

def list_parser():

    wb = openpyxl.reader.excel.load_workbook(filename="PRICE2022.xlsx", data_only=True)  # открываем файл с переменными

    list_par = wb.sheetnames # список листов

    return (list_par)

    # for sheet in wb: # печатаем список листов по одному
    #     print(sheet.title)






root = Tk() # класс окна

root.title("ПАРСЕР ЛКМ SYMPHONY") #текст в шапке окна
root.geometry("500x500") # размер окна

photo = tkinter.PhotoImage(file='logofkm.png') # лого в шапке окна
root.iconphoto(False,photo)

download_label = Label(root, text = 'Выберите что будем проверять', font=35) # текст в окне
download_label.pack(pady = 20)

def checkbutton_changed():

    showinfo(title="Info",message=enabled.get())


# создаем чекбоксы

list_par=list_parser() #обращаемся к функции для получения списка листов для парсинга

enabled = StringVar()
n=0

for i in list_par:
    if n != 0: # Пропускаем первый лист где находится контрольные цены
        enabled_checkbutton = tkinter.Checkbutton(root, text=i, variable=enabled, offvalue="", onvalue=i, command=checkbutton_changed)
        enabled_checkbutton.pack(padx=6, pady=6, anchor=NW)
    n = +1
#
# prs2 = tkinter.Checkbutton(root, text=list_par[2])
# prs2.pack()
# prs3 = tkinter.Checkbutton(root, text=list_par[3])
# prs3.pack()



button = Button(root,text=" Запустить Парсер ", font = 40, command= list_parser) # кнопка
button.pack(side = BOTTOM, pady = 40) # расположение кнопки, отступ по оси У

root.mainloop() # запуск визуализации окна


#

# wb.active = 1 # делаем активной вторую страницу  там где Бафус
# sheet_parser = wb.active # копируем страницу в переменную
# wb.active = 0 # делаем активной первую страницу с прайсом РРЦ
# sheet_active = wb.active
