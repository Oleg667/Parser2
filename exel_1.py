import openpyxl
from openpyxl.utils import get_column_letter
import re

path_to_file = 'PRICE2022.xlsx'

# search_text = input(str('Какой текст ищем: '))
# search_text = search_text.lower()
search_text = '00-00009567'
print('Ищем:', search_text)

wb = openpyxl.load_workbook(path_to_file)  # Грузим наш прайс-лист
sheets_list = wb.sheetnames  # Получаем список всех листов в файле
sheet_active = wb[sheets_list[0]]  # Начинаем работать с самым первым
row_max = sheet_active.max_row  # Получаем количество столбцов
# print(type(row_max))
column_max = sheet_active.max_column  # Получаем количество строк

print('В файле:', path_to_file, '\n Строк:', row_max, '\n Столбцов:', column_max)

row_min = 1  # Переменная, отвечающая за номер строки
column_min = 1  # Переменная, отвечающая за номер столбца

while column_min <= column_max:
    row_min_min = row_min
    row_max_max = row_max
    while row_min_min <= row_max_max:
        row_min_min = str(row_min_min)

        word_column = get_column_letter(column_min)
        word_column = str(word_column)
        word_cell = word_column + row_min_min

        data_from_cell = sheet_active[word_cell].value
        data_from_cell = str(data_from_cell)
        # print(data_from_cell)
        regular = search_text
        result = re.findall(regular, data_from_cell)
        if len(result) > 0:
            print('Нашли в ячейке:', word_cell)
            RRC='C' + row_min_min
            print( 'РРЦ равняется ',sheet_active[RRC].value,' рубль' )
        row_min_min = int(row_min_min)
        row_min_min = row_min_min + 1
    column_min = column_min + 1