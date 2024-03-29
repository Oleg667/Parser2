import datetime
import chromedriver_binary

import pandas as pd
import json

import openpyxl
import time
import re



from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl.utils import get_column_letter # Функция get_column_letter() преобразовывает индекс столбца idx в букву столбца (3 => C)

# #  Создаем браузер
# browser = webdriver.Chrome()
# browser.maximize_window()




# Парсинг цены. Входные данные:
# - адрес страницы
# - локатор цены
# - локатор наименования лкм
# - локатор тары
# - локатор базы

def open_get(cod_sku,url_1,xpath_prise,xpath_lkm,xpath_tara,xpath_baza): #открываем страницу по адресу и вытаскивает цену

        browser.get(url_1) #открытие страницы по переданному адресу

        time.sleep(0.7)

        err='Цена' # err - переменная показывающая какой этап парсинга выполняется, для идентификация ошибки
        price_n=' '
        try:
                price = browser.find_element(By.XPATH, xpath_prise) # получение цены
                price_n = price.text # форматирование данных цены
        except NoSuchElementException:  # ошибка, из-за которой этот код работает не так, как ожидалось
                print( cod_sku,"неверный XPATH для  ",err)
                pass
        except StaleElementReferenceException:  # ошибка, из-за которой этот код работает не так, как ожидалось
                print( cod_sku,"элемент закрыт'  ",err)
                pass
        err = 'Имя SKU'
        name_n = ' '
        try:
                name = browser.find_element(By.XPATH, xpath_lkm)  # получение наименования
                name_n = name.text  # форматирование данных наименование
        except NoSuchElementException:  # spelling error making this code not work as expected
                print( cod_sku,"неверный XPATH для  ",err)
                pass
        except StaleElementReferenceException:  # ошибка, из-за которой этот код работает не так, как ожидалось
                print( cod_sku,"элемент закрыт'  ",err)
                pass

        err = 'Тара'
        tara_n = ' '
        try:
                tara = browser.find_element(By.XPATH, xpath_tara)       # получение тары
                tara_n = tara.text                                      # форматирование данных тары
        except NoSuchElementException:                                  # spelling error making this code not work as expected
                print( cod_sku,"неверный XPATH для  ",err)
                pass
        except StaleElementReferenceException:  # ошибка, из-за которой этот код работает не так, как ожидалось
                print( cod_sku,"элемент закрыт'  ",err)
                pass

        err = 'База'
        baza_n = ' '
        try:
                baza = browser.find_element(By.XPATH, xpath_baza)  # получение базы
                baza_n = baza.text  # форматирование данных базы
        except NoSuchElementException:  # spelling error making this code not work as expected
                print( cod_sku,"неверный XPATH для  ",err)
                pass
        except StaleElementReferenceException:  # ошибка, из-за которой этот код работает не так, как ожидалось
                print( cod_sku,"элемент закрыт'  ",err)
                pass
        # browser.close() #закрываем браузер

        # time.sleep(3)


        return(name_n,tara_n,baza_n,price_n)


def price_rrc(cod_sku,sheet_active): #получаем РРЦ из прайс-листа по коду товара

        path_to_file = 'PRICE2022.xlsx'

        search_text = cod_sku

        wb = openpyxl.load_workbook(path_to_file)  # Грузим наш прайс-лист

        sheets_list = wb.sheetnames  # Получаем список всех листов в файле

        row_max = sheet_active.max_row  # Получаем количество столбцов

        column_max = sheet_active.max_column  # Получаем количество строк

        row_min = 1  # Переменная, отвечающая за номер строки
        column_min = 1  # Переменная, отвечающая за номер столбца

        while column_min <= column_max:
                row_min_min = row_min
                row_max_max = row_max
                while row_min_min <= row_max_max:
                        row_min_min = str(row_min_min)

                        word_column = get_column_letter(column_min)
                        word_column = str(word_column)
                        word_cell = word_column + row_min_min

                        data_from_cell = sheet_active[word_cell].value
                        data_from_cell = str(data_from_cell)
                        # print(data_from_cell)
                        regular = search_text
                        result = re.findall(regular, data_from_cell)
                        if len(result) > 0:

                                RRC = 'C' + row_min_min
                                RRC_SKU = sheet_active[RRC].value # переменная хранит найденное РРЦ

                        row_min_min = int(row_min_min)
                        row_min_min = row_min_min + 1
                column_min = column_min + 1
        return(RRC_SKU)



def Parser_master(selected_langs1,skidka_5,violation): # Основная функция парсинга цен

        #  Создаем браузер
        global browser # переменная хранит браузер
        browser = webdriver.Chrome()
        browser.maximize_window()



        wb = openpyxl.reader.excel.load_workbook(filename="PRICE2022.xlsx", data_only=True) #открываем файл с переменными

        wb.active = 0  # делаем активной первую страницу с прайсом РРЦ
        sheet_active = wb.active

        # Создаем заголовок таблицы ... не актуально.
        # data={ 'Cod': ['1'],
        #        'NAME SKU': ['2'],
        #        'Prise in website': ['3'],
        #        'recommended retail price list': ['4'],
        #        'comparison': ['5'],
        #        'URL': ['6']  }

        # df = pd.DataFrame(data) #Создаем двумерную таблицу отчета

        # Основной цикл проверки, открываем листы по очереди и начиная со второй строки пока не будет прочитана команда stop в первом столбце
        print(selected_langs1)

        df = pd.DataFrame(
                {'Cod': [1], 'NAME SKU': [2], 'Prise in website': [3], 'recommended retail price list': [4],
                 'comparison': [5], 'URL': [5]})  # Создаем заголовки столбцов в файле отчета

        for sheet_name in selected_langs1: # запускаем парсинг-проверку выбранных интернет магазинов (по названию страниц в файле эксель)
                print(sheet_name,skidka_5,violation)
                wb.active = wb.get_sheet_by_name(sheet_name)  # делаем активной очередную страницу из выбранного списка
                sheet_parser = wb.active  # копируем страницу в переменную

                i=2 # начинаем просмотр данных для поиска со второй строки в файле PRICE2022

                while i>0:

                        # извлекаем данные для парсинга из файла PRICE2022

                        cod_sku = sheet_parser['a'+ str(i)].value
                        url_1= sheet_parser['b'+ str(i)].value
                        xpath_prise = sheet_parser['c'+ str(i)].value
                        xpath_lkm = sheet_parser['d'+ str(i)].value
                        xpath_tara = sheet_parser['e'+ str(i)].value
                        xpath_baza = sheet_parser['f'+ str(i)].value

                        if cod_sku == "stop": # если прочитали в первом столбце 'stop' то переходим к следующему листу или заканчиваем проверку если это был последний лист

                                break

                        parser = open_get(cod_sku,url_1,xpath_prise,xpath_lkm,xpath_tara,xpath_baza) # получаем данные по продукту с сайта


                        Price_parser =int(''.join([i for i in parser[3] if i.isdigit()])) # удаляем все символы кроме цифр, и преобразуем в число
                        Lkm = parser[0]
                        Tara = parser[1]
                        Baza = parser[2]
                        Sku_rrc = int(price_rrc(cod_sku,sheet_active))

                        Name_sku = Lkm+' '+Tara+' '+Baza
                        print(Sku_rrc,skidka_5,Price_parser)

                        if Sku_rrc*skidka_5 > Price_parser or violation == 0: # Если цена на сайте меньше чем контрольная цена (или контрольная цена -5%), то добавляем информацию о нарушении

                                # Гененрируем  новую строку

                                comparison = (Sku_rrc-Price_parser)/Sku_rrc*100 # % отличия цены на сайте от контрольной цены


                                new_row = {'Cod': cod_sku, 'NAME SKU': Name_sku, 'Prise in website': Price_parser,
                                         'recommended retail price list': Sku_rrc, 'comparison': comparison,
                                         'URL': url_1}

                                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True) # добавляем строку в таблицу отчета



                        i = i+1 # номер строки для выборки увеличен на 1

        now = datetime.datetime.now()
        report ="./"+str(now.year) + str(now.month) + str(now.day) + str(now.hour) + str(now.minute) + str(now.second) + ".xlsx"
        #df.to_excel(report, index=False)  # Записываем файл с созданными значениями и привязкой к дате и времени
        df.to_excel('./report.xlsx', index=False) # Записываем файл с созданными значениями
        browser.close()  # закрываем браузер


